# Copyright 2024, Polytechnique Montreal and contributors
# This file is licensed under the MIT License.
# License text available at https://opensource.org/licenses/MIT

# Note: This script includes functions that help generate and test Generator scripts.
import os  # File system operations
import openpyxl  # Read data from Excel
from openpyxl import Workbook  # Read data from Excel, File system operations
from typing import List, Union  # Types for Python

# Define constants
MOCKER_EXCEL_FILE = "generator/examples/test.xlsx"
INDENT = "    "  # 4-space indentation


# Add Generator comment at the start of the file
def add_generator_comment() -> str:
    ts_code = f"// This file was automatically generated by the Evolution Generator.\n"
    ts_code += f"// The Evolution Generator is used to automate the creation of consistent, reliable code.\n"
    ts_code += f"// Any changes made to this file will be overwritten.\n\n"
    return ts_code


# TODO: Add types for rows and headers
# Read data from Excel and return rows and headers
def get_data_from_excel(input_file: str, sheet_name: str) -> tuple:
    try:
        # Load Excel file
        workbook: Workbook = openpyxl.load_workbook(input_file, data_only=True)
        sheet = workbook[sheet_name]  # Get InputRange sheet
        rows = list(sheet.rows)  # Get all rows in the sheet
        headers = [cell.value for cell in rows[0]]  # Get headers from the first row

        # Error when header has spaces
        if any(" " in header for header in headers):
            raise Exception("Header has spaces")

        # Error when header is None
        if None in headers:
            raise Exception("Header is None")

        return rows, headers

    except Exception as e:
        print(f"Error reading Excel in {sheet_name} sheet: {e}")
        raise e


# TODO: Add types for rows and headers
# Get values from the row
def get_values_from_row(row, headers) -> tuple:
    try:
        # Create a dictionary from the row values and headers
        row_dict = dict(zip(headers, (cell.value for cell in row)))
        values = []  # List of values from the row

        # Get values from the row dictionary
        for header in headers:
            header = row_dict[header]
            values.append(header)

        return values

    except Exception as e:
        print(f"Error getting values from row: {e}")
        raise e


# Error when any required fields values are None
def error_when_missing_required_fields(
    required_fields_names, required_fields_values, row_number: int
):
    if any(field is None for field in required_fields_values):
        missing_fields = [
            field_name
            for field_value, field_name in zip(
                required_fields_values, required_fields_names
            )
            if field_value is None
        ]
        raise Exception(
            f"Required field is missing in row {row_number}. Missing fields: {missing_fields}"
        )


# Generate output file
def generate_output_file(ts_code: str, output_file: str):
    try:
        with open(output_file, mode="w", encoding="utf-8", newline="\n") as ts_file:
            ts_file.write(ts_code)

        print(f"Generate {output_file} successfully")

    except Exception as e:
        print(f"Error generating {output_file}: {e}")
        raise e


# Create mocked Excel data for testing
def create_mocked_excel_data(
    sheet_name: str, headers: List[str], rows_data: List[List[Union[str, int, float]]]
) -> Workbook:
    workbook: Workbook = openpyxl.Workbook()  # Create a workbook
    sheet = workbook.active  # Get the active sheet
    sheet.title = sheet_name  # Change sheet title

    # Add headers
    sheet.append(headers)

    # Iterate through each row data
    for row_data in rows_data:
        sheet.append(row_data)  # Add row data

    # Create the excel file
    workbook.save(MOCKER_EXCEL_FILE)

    # Return the workbook
    return workbook


# Delete file if exists
def delete_file_if_exists(file_path: str) -> None:
    if os.path.isfile(file_path):
        os.remove(file_path)


# Check if the input file is an Excel file
def is_excel_file(file: str) -> None:
    if not file.endswith(".xlsx"):
        raise Exception(
            f"Invalid input file extension for {file} : must be an Excel .xlsx file"
        )


# Check if the output file is an TypeScript file
def is_ts_file(file: str) -> None:
    if not file.endswith(".tsx"):
        raise Exception(
            f"Invalid output file extension for {file} : must be an TypeScript .tsx file"
        )


# Check if the sheet exists
def sheet_exists(workbook: Workbook, sheet_name: str) -> None:
    if sheet_name not in workbook.sheetnames:
        raise Exception(f"Invalid sheet name in {sheet_name} sheet")


# Get workbook from Excel file
def get_workbook(input_file: str) -> Workbook:
    workbook = openpyxl.load_workbook(input_file, data_only=True)
    return workbook


# Get headers from the first row
def get_headers(sheet, expected_headers: List[str], sheet_name: str) -> List[str]:
    # Get headers from the first row
    current_headers = [cell.value for cell in list(sheet.rows)[0]]

    # Check if the good numbers of headers
    if len(current_headers) != len(expected_headers):
        raise Exception(f"Invalid number of column in {sheet_name} sheet")

    # Check if the headers are valid
    if current_headers != expected_headers:
        raise Exception(f"Invalid headers in {sheet_name} sheet")

    return current_headers